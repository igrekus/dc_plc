# -*- coding: utf-8 -*-

from __future__ import unicode_literals

import openpyxl
import frappe
import datetime
import zipfile
import os
import pdfkit

from six import BytesIO
from jinja2 import Environment, FileSystemLoader

small_fields = {
	'ext_num': u'Отраслевой номер',
	'rnd_proj': u'ОКР',
	'int_num': u'Внутренний номер',
	'type': u'Тип',
	'func': u'Функция',
	'letter': 'uЛитерность',
	'opcon_num': u'Номер ТУ',
	'status': u'Развитие',
	'devs': u'Разработчик',
	'cons': u'Консультант',
	'chip': u'Кристалл',
	'package': u'Корпус',
	'desdoc_num': u'Номер КД',
	'asm_board': u'Плата в сборке',
	'procmap_num': u'Номер ТК',
}


large_fields = {
	'application': u'Применение',
	'analogs': u'Аналоги',
	'final_description': u'Финально описание',
	'description': u'Техническое описание',
	'specs': u'Параметры',
	'reports': u'Отчёты',
	'datasheet': u'Даташит',
}


class ExcelProductListExport:
	def __init__(self, headers, fields, data):
		self._headers = headers
		self._rows = [[data_object[f] for f in fields] for data_object in data]

	def _append_header(self, sheet):
		header_cell = sheet['A1']
		for col, header in enumerate(['№'] + self._headers):
			header_cell.offset(0, 0 + col).value = header

	def _append_data(self, sheet):
		data_cell = sheet['A2']
		for row, row_data in enumerate(self._rows):
			for col, value in enumerate(row_data):
				data_cell.offset(0 + row, 0).value = row + 1
				data_cell.offset(0 + row, 1 + col).value = value

	@property
	def name(self):
		return 'Product list {}.xlsx'.format(datetime.date.today().isoformat())

	@property
	def bytes(self):
		wb = openpyxl.Workbook()
		ws = wb.active

		self._append_header(ws)
		self._append_data(ws)

		xlsx_file = BytesIO()
		wb.save(xlsx_file)
		return xlsx_file.getvalue()


class ExcelProductCardExport:

	def __init__(self, headers, fields, product_data):
		self._headers = headers
		self._fields = {f: h for f, h in zip(fields, headers)}
		self._data = {field: product_data[field] or '-' for field in fields}
		self._small_fields = [field for field in small_fields.keys() if field in self._fields]
		self._large_fields = [field for field in large_fields.keys() if field in self._fields]

	def _append_header(self, sheet):
		header_cell = sheet['A1']
		header_cell.value = 'Карточка изделия'

	def _append_small_fields(self, sheet):
		base_cell = sheet['A2']
		col = 0
		row = 0
		for field in self._small_fields:
			base_cell.offset(0 + row, 0 + col * 3).value = self._fields[field]
			base_cell.offset(0 + row, 1 + col * 3).value = self._data[field]
			col += 1
			if col > 1:
				col = 0
				row += 1

	def _append_large_fields(self, sheet):
		base_cell = sheet['A' + str(sheet.max_row + 2)]
		for row, field in enumerate(self._large_fields):
			base_cell.offset(0 + row, 0). value = self._fields[field]
			base_cell.offset(0 + row, 1). value = self._data[field]

	@property
	def name(self):
		return '{} {} {}.xlsx'.format(
			self._data.get('int_num', '-'),
			self._data.get('ext_num', '-'),
			datetime.date.today().isoformat()
		)

	@property
	def bytes(self):
		wb = openpyxl.Workbook()
		ws = wb.active

		self._append_header(ws)
		self._append_small_fields(ws)
		self._append_large_fields(ws)

		xlsx_file = BytesIO()
		wb.save(xlsx_file)

		return xlsx_file.getvalue()


def export_xlsx(exports, headers, fields, data):
	zipname = './site1.local/public/files/outzip-' + frappe.generate_hash('', 10) + '.zip'
	with zipfile.ZipFile(zipname, 'w') as outzip:
		if exports['list']:
			export_list = ExcelProductListExport(headers, fields, data)
			outzip.writestr(export_list.name, export_list.bytes)

		if exports['cards']:
			for prod in data:
				export = ExcelProductCardExport(headers, fields, prod)
				outzip.writestr(export.name, export.bytes)

		if exports['datasheets']:
			# TODO append datasheets to zip
			pass

	with open(zipname, mode='rb') as f:
		frappe.response['filename'] = 'product_export.zip'
		frappe.response['filecontent'] = BytesIO(f.read()).getvalue()
		frappe.response['type'] = 'binary'

	os.remove(zipname)


class ProductPDF:
	def __init__(self, exports, headers, fields, product_data):
		self._filename = f'./site1.local/public/files/outpdf-{frappe.generate_hash("", 10)}.pdf'
		self._headers = headers
		self._fields = {f: h for f, h in zip(fields, headers)}
		self._products = product_data
		self._exports = exports

		self._small_fields = [field for field in small_fields.keys() if field in self._fields]
		self._large_fields = [field for field in large_fields.keys() if field in self._fields]

	@property
	def html(self):
		env = Environment(loader=FileSystemLoader('./assets/dc_plc/html/'))
		template = env.get_template('export_template.html')
		return template.render(params={
			'has_list': self._exports['list'],
			'has_cards': self._exports['cards'],
			'headers': self._headers,
			'list': self.list_data,
			'cards': self.card_data,
		})

	@property
	def card_data(self):
		cards = list()
		for prod in self._products:
			col = 0
			row_data = list()
			small_rows = list()
			for field in self._small_fields:
				row_data.append((self._fields[field], prod[field] or '-'))
				col += 1
				if col > 1:
					col = 0
					small_rows.append(row_data)
					row_data = list()

			large_rows = [(self._fields[field], prod[field] or '-') for field in self._large_fields]

			cards.append({
				'small_rows': small_rows,
				'large_rows': large_rows
			})
		return cards

	@property
	def list_data(self):
		rows = [[product[f] or '-' for f in self._fields] for product in self._products]
		return rows

	@property
	def name(self):
		return self._filename

	@property
	def bytes(self):
		filedata = pdfkit.from_string(self.html, False, options={})
		return BytesIO(filedata).getvalue()


def export_pdf(exports, headers, fields, product_data):
	pdf = ProductPDF(exports, headers, fields, product_data)
	return pdf.html
	# frappe.local.response.filename = "product_export.html"
	# frappe.local.response.filecontent = pdf.html
	# frappe.local.response.type = "html"
